from openpyxl import load_workbook

def headers(ws,row):
    return [c.value if c.value is not None else '' for c in ws[row]]

def validate_file(ref,target,cfg):
    out=[]
    refwb=load_workbook(ref,read_only=True,data_only=True)
    tgtwb=load_workbook(target,read_only=True,data_only=True)

    for sheet,row in cfg.items():
        if sheet not in tgtwb.sheetnames:
            out.append(('FAIL',sheet,f'Required worksheet is missing: {sheet}'))
            continue

        rh=headers(refwb[sheet],row)
        th=headers(tgtwb[sheet],row)

        blanks=[str(i+1) for i,v in enumerate(th) if str(v).strip()=='']
        if blanks:
            out.append(('FAIL',sheet,'Blank header column(s) detected at column position(s): '+','.join(blanks)))
            continue

        if rh!=th:
            missing=[x for x in rh if x not in th]
            extra=[x for x in th if x not in rh]
            msg=[]
            if missing:
                msg.append('Required column(s) are missing: '+', '.join(map(str,missing)))
            if extra:
                msg.append('Unexpected additional column(s) detected: '+', '.join(map(str,extra)))
            if not msg:
                for i,(a,b) in enumerate(zip(rh,th),1):
                    if a!=b:
                        msg.append(f"Header name mismatch detected at column position {i}. Expected '{a}' but found '{b}'.")
                        break
            out.append(('FAIL',sheet,' | '.join(msg)))
        else:
            out.append(('PASS',sheet,'All sheet names and headers passed validation.'))
    return out
