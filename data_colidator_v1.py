"""
Excel Data Consolidator — GUI Tool
Combines data from multiple Excel files into one or more output files,
splitting when a row-count limit is reached. Supports multi-row headers.
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import glob
import math
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import copy


# ─────────────────────────────────────────────────────────────────
# Core consolidation logic
# ─────────────────────────────────────────────────────────────────

def read_sheet_data(filepath, sheet_index, header_row_count):
    """
    Read header rows and data rows separately from one sheet.
    Returns:
        headers : list of tuples — one tuple per header row,
                  each tuple is the raw cell values for that row
        data    : list of tuples — data rows (after the header rows)
        col_count: int — max column count seen across all rows
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    if sheet_index >= len(sheet_names):
        wb.close()
        raise ValueError(
            f"Sheet at position {sheet_index + 1} does not exist in '{os.path.basename(filepath)}'."
        )
    ws = wb.worksheets[sheet_index]

    headers = []
    data = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < header_row_count:
            headers.append(tuple(row))
        else:
            # Skip completely empty rows
            if any(c is not None and str(c).strip() != "" for c in row):
                data.append(tuple(row))

    col_count = 0
    all_rows = headers + data
    if all_rows:
        col_count = max(len(r) for r in all_rows)

    wb.close()
    return headers, data, col_count


def consolidate(file_paths, sheet_index, header_row_count,
                output_basename, output_dir, row_limit, status_cb=None):
    """
    Consolidate data from all files into one or more output Excel files.

    Parameters
    ----------
    file_paths      : list of str — input file paths
    sheet_index     : int — 0-based sheet position to read from each file
    header_row_count: int — number of header rows in each file
    output_basename : str — base name for output files (no extension)
    output_dir      : str — directory to save output files
    row_limit       : int or None — max data rows per output file (None = unlimited)
    status_cb       : callable(msg) — optional callback for progress messages

    Returns
    -------
    list of str — paths of output files created
    summary dict  — {total_files, total_data_rows, output_files, errors}
    """

    def log(msg):
        if status_cb:
            status_cb(msg)

    errors = []
    all_data_rows = []
    reference_headers = None
    reference_col_count = 0

    # ── Pass 1: collect all data ─────────────────────────────────
    for fp in file_paths:
        fname = os.path.basename(fp)
        try:
            hdrs, data, col_count = read_sheet_data(fp, sheet_index, header_row_count)
            if reference_headers is None:
                reference_headers = hdrs
                reference_col_count = col_count
            else:
                # Keep the widest column count seen
                reference_col_count = max(reference_col_count, col_count)
            all_data_rows.extend(data)
            log(f"Read {len(data):,} data row(s) from {fname}")
        except Exception as e:
            errors.append({"file": fname, "error": str(e)})
            log(f"  ⚠ Skipped {fname}: {e}")

    if not all_data_rows:
        raise RuntimeError("No data rows were collected from the selected files.")

    if reference_headers is None:
        reference_headers = []

    total_data_rows = len(all_data_rows)

    # ── Pass 2: split into chunks ─────────────────────────────────
    if row_limit and row_limit > 0:
        chunks = [
            all_data_rows[i: i + row_limit]
            for i in range(0, total_data_rows, row_limit)
        ]
    else:
        chunks = [all_data_rows]

    output_files = []
    num_files = len(chunks)
    pad = len(str(num_files))  # zero-pad width

    for file_idx, chunk in enumerate(chunks):
        if num_files == 1:
            out_name = f"{output_basename}.xlsx"
        else:
            out_name = f"{output_basename}_{str(file_idx + 1).zfill(pad)}.xlsx"

        out_path = os.path.join(output_dir, out_name)
        _write_output(out_path, reference_headers, chunk,
                      reference_col_count, file_idx + 1, num_files)
        output_files.append(out_path)
        log(f"Saved {out_name}  ({len(chunk):,} data rows)")

    summary = {
        "total_input_files": len(file_paths),
        "files_read": len(file_paths) - len(errors),
        "total_data_rows": total_data_rows,
        "output_files": output_files,
        "errors": errors,
    }
    return output_files, summary


def _write_output(out_path, headers, data_rows, col_count, file_num, total_files):
    """Write one consolidated output file with styled headers."""
    wb = openpyxl.Workbook(write_only=False)
    ws = wb.active
    ws.title = "Consolidated Data"

    # Styles
    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    hdr_fill = PatternFill("solid", fgColor="1F3864")   # dark navy
    alt_fill = PatternFill("solid", fgColor="EEF2F7")   # light blue-grey
    thin = Side(style="thin", color="C0C8D8")
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_al = Alignment(horizontal="left", vertical="center", wrap_text=False)

    # ── Write header rows ────────────────────────────────────────
    header_row_count = len(headers)
    for hi, hrow in enumerate(headers):
        row_out = []
        for ci in range(col_count):
            val = hrow[ci] if ci < len(hrow) else None
            row_out.append(val)
        ws.append(row_out)
        excel_row = hi + 1
        for ci in range(1, col_count + 1):
            cell = ws.cell(row=excel_row, column=ci)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.border = cell_border
            cell.alignment = center
        ws.row_dimensions[excel_row].height = 18

    # ── Write data rows ──────────────────────────────────────────
    for di, drow in enumerate(data_rows):
        row_out = []
        for ci in range(col_count):
            val = drow[ci] if ci < len(drow) else None
            row_out.append(val)
        ws.append(row_out)
        excel_row = header_row_count + di + 1
        fill = alt_fill if di % 2 == 1 else None
        for ci in range(1, col_count + 1):
            cell = ws.cell(row=excel_row, column=ci)
            if fill:
                cell.fill = fill
            cell.border = cell_border
            cell.alignment = left_al
            cell.font = Font(name="Arial", size=9)

    # ── Column widths (auto-fit based on header names) ───────────
    for ci in range(1, col_count + 1):
        max_len = 8
        for hi in range(header_row_count):
            cell = ws.cell(row=hi + 1, column=ci)
            if cell.value:
                max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len, 40)

    # ── Freeze header rows ───────────────────────────────────────
    if header_row_count > 0:
        ws.freeze_panes = ws.cell(row=header_row_count + 1, column=1)

    # ── Meta info in a second sheet ──────────────────────────────
    ws_info = wb.create_sheet("Consolidation Info")
    ws_info.column_dimensions["A"].width = 28
    ws_info.column_dimensions["B"].width = 50
    meta = [
        ("Generated on", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Output file", f"{file_num} of {total_files}"),
        ("Data rows in this file", len(data_rows)),
        ("Header rows", header_row_count),
        ("Columns", col_count),
    ]
    for r, (k, v) in enumerate(meta, 1):
        ws_info.cell(row=r, column=1, value=k).font = Font(bold=True, name="Arial", size=10)
        ws_info.cell(row=r, column=2, value=str(v)).font = Font(name="Arial", size=10)

    wb.save(out_path)


# ─────────────────────────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Data Consolidator")
        self.geometry("1080x760")
        self.minsize(880, 620)
        self.configure(bg="#0f1117")

        self.input_mode = tk.StringVar(value="files")
        self.input_paths = []          # file paths when mode=files
        self.input_folder = tk.StringVar()
        self.sheet_position = tk.StringVar(value="1")
        self.header_rows = tk.StringVar(value="1")
        self.output_name = tk.StringVar(value="consolidated")
        self.output_dir = tk.StringVar()
        self.row_limit = tk.StringVar(value="")   # blank = no limit
        self.use_limit = tk.BooleanVar(value=False)

        self._build_styles()
        self._build_ui()

    # ── Styles ────────────────────────────────────────────────────
    def _build_styles(self):
        s = ttk.Style(self)
        s.theme_use("clam")
        BG, CARD = "#0f1117", "#1a1d27"
        ACCENT, FG, MUTED = "#4f8ef7", "#e8eaf0", "#6b7280"
        s.configure("TFrame", background=BG)
        s.configure("Card.TFrame", background=CARD)
        s.configure("TLabel", background=BG, foreground=FG, font=("Arial", 10))
        s.configure("Card.TLabel", background=CARD, foreground=FG, font=("Arial", 10))
        s.configure("Head.TLabel", background=CARD, foreground=ACCENT,
                    font=("Arial", 11, "bold"))
        s.configure("Muted.TLabel", background=CARD, foreground=MUTED,
                    font=("Arial", 8))
        s.configure("Title.TLabel", background=BG, foreground=FG,
                    font=("Arial", 18, "bold"))
        s.configure("TButton", background=ACCENT, foreground="#ffffff",
                    font=("Arial", 10, "bold"), borderwidth=0, padding=6)
        s.map("TButton",
              background=[("active", "#6ba3ff"), ("disabled", "#2d3148")],
              foreground=[("disabled", MUTED)])
        s.configure("Ghost.TButton", background=CARD, foreground=FG,
                    font=("Arial", 10), borderwidth=1, padding=5)
        s.map("Ghost.TButton", background=[("active", "#252839")])
        s.configure("Run.TButton", background="#22c55e", foreground="#000000",
                    font=("Arial", 11, "bold"), padding=8)
        s.map("Run.TButton", background=[("active", "#16a34a")])
        s.configure("TEntry", fieldbackground="#252839", foreground=FG,
                    insertcolor=FG, borderwidth=0)
        s.configure("TCheckbutton", background=CARD, foreground=FG,
                    font=("Arial", 10))
        s.configure("TRadiobutton", background=BG, foreground=FG, font=("Arial", 10))
        s.configure("Treeview", background=CARD, fieldbackground=CARD,
                    foreground=FG, font=("Arial", 9), rowheight=26, borderwidth=0)
        s.configure("Treeview.Heading", background="#252839", foreground=ACCENT,
                    font=("Arial", 10, "bold"), borderwidth=0)
        s.map("Treeview", background=[("selected", "#2d3f6b")])
        self.C = {"bg": BG, "card": CARD, "accent": ACCENT,
                  "fg": FG, "muted": MUTED,
                  "pass": "#22c55e", "fail": "#ef4444", "warn": "#f59e0b"}

    # ── Top-level layout ──────────────────────────────────────────
    def _build_ui(self):
        # Header bar
        bar = tk.Frame(self, bg="#0d0f18", height=54)
        bar.pack(fill="x")
        bar.pack_propagate(False)
        tk.Label(bar, text="⬡  Excel Data Consolidator",
                 bg="#0d0f18", fg=self.C["accent"],
                 font=("Arial", 15, "bold")).pack(side="left", padx=20, pady=14)
        tk.Label(bar, text="Merge data from multiple files, split by row limit",
                 bg="#0d0f18", fg=self.C["muted"],
                 font=("Arial", 9)).pack(side="left", pady=14)

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=14, pady=10)

        left = ttk.Frame(body, width=330)
        left.pack(side="left", fill="y", padx=(0, 10))
        left.pack_propagate(False)
        self._build_left(left)

        right = ttk.Frame(body)
        right.pack(side="left", fill="both", expand=True)
        self._build_right(right)

    # ── Left panel ────────────────────────────────────────────────
    def _build_left(self, parent):
        def card(title):
            f = ttk.Frame(parent, style="Card.TFrame", padding=12)
            f.pack(fill="x", pady=(0, 10))
            ttk.Label(f, text=title, style="Head.TLabel").pack(anchor="w", pady=(0, 8))
            return f

        # ① Input source
        c1 = card("① Input Source")
        rb = ttk.Frame(c1, style="Card.TFrame")
        rb.pack(fill="x", pady=(0, 6))
        ttk.Radiobutton(rb, text="Select Files", variable=self.input_mode,
                        value="files", command=self._toggle_input).pack(side="left", padx=(0, 14))
        ttk.Radiobutton(rb, text="Select Folder", variable=self.input_mode,
                        value="folder", command=self._toggle_input).pack(side="left")

        # Files list
        self._files_frame = ttk.Frame(c1, style="Card.TFrame")
        self._files_frame.pack(fill="x")
        lf = tk.Frame(self._files_frame, bg=self.C["card"])
        lf.pack(fill="x")
        sb = ttk.Scrollbar(lf)
        sb.pack(side="right", fill="y")
        self.files_lb = tk.Listbox(lf, height=5, bg="#252839", fg=self.C["fg"],
                                   selectbackground=self.C["accent"],
                                   font=("Arial", 8), relief="flat",
                                   yscrollcommand=sb.set, borderwidth=0,
                                   highlightthickness=0)
        self.files_lb.pack(fill="x")
        sb.config(command=self.files_lb.yview)
        br = ttk.Frame(self._files_frame, style="Card.TFrame")
        br.pack(fill="x", pady=(4, 0))
        ttk.Button(br, text="+ Add Files", style="Ghost.TButton",
                   command=self._add_files).pack(side="left")
        ttk.Button(br, text="Remove Selected", style="Ghost.TButton",
                   command=self._remove_selected).pack(side="left", padx=5)
        ttk.Button(br, text="Clear All", style="Ghost.TButton",
                   command=self._clear_files).pack(side="left")

        # Folder picker
        self._folder_frame = ttk.Frame(c1, style="Card.TFrame")
        fr = ttk.Frame(self._folder_frame, style="Card.TFrame")
        fr.pack(fill="x")
        ttk.Entry(fr, textvariable=self.input_folder,
                  style="TEntry").pack(side="left", fill="x", expand=True, ipady=4)
        ttk.Button(fr, text="Browse", style="Ghost.TButton",
                   command=self._browse_folder).pack(side="left", padx=(5, 0))
        ttk.Label(self._folder_frame,
                  text="All .xlsx / .xlsm / .xls files in the folder will be included.",
                  style="Muted.TLabel", wraplength=280).pack(anchor="w", pady=(4, 0))

        self._toggle_input()

        # ② Sheet & header config
        c2 = card("② Sheet & Header Configuration")
        r1 = ttk.Frame(c2, style="Card.TFrame")
        r1.pack(fill="x", pady=(0, 6))
        ttk.Label(r1, text="Sheet position (1 = first):",
                  style="Card.TLabel").pack(side="left")
        ttk.Entry(r1, textvariable=self.sheet_position, width=5,
                  style="TEntry").pack(side="left", padx=(6, 0), ipady=3)

        r2 = ttk.Frame(c2, style="Card.TFrame")
        r2.pack(fill="x")
        ttk.Label(r2, text="Header row count:", style="Card.TLabel").pack(side="left")
        ttk.Entry(r2, textvariable=self.header_rows, width=5,
                  style="TEntry").pack(side="left", padx=(6, 0), ipady=3)
        ttk.Label(c2,
                  text="Header rows are taken from the first file and repeated in every output file.",
                  style="Muted.TLabel", wraplength=280).pack(anchor="w", pady=(6, 0))

        # ③ Output settings
        c3 = card("③ Output Settings")
        r3 = ttk.Frame(c3, style="Card.TFrame")
        r3.pack(fill="x", pady=(0, 6))
        ttk.Label(r3, text="Output file name:", style="Card.TLabel").pack(side="left")
        ttk.Entry(r3, textvariable=self.output_name, width=14,
                  style="TEntry").pack(side="left", padx=(6, 0), ipady=3)

        r4 = ttk.Frame(c3, style="Card.TFrame")
        r4.pack(fill="x", pady=(0, 6))
        ttk.Label(r4, text="Output folder:", style="Card.TLabel").pack(side="left")
        ttk.Entry(r4, textvariable=self.output_dir, width=10,
                  style="TEntry").pack(side="left", padx=(6, 0), fill="x", expand=True, ipady=3)
        ttk.Button(r4, text="…", style="Ghost.TButton", width=3,
                   command=self._browse_output_dir).pack(side="left", padx=(4, 0))

        # Row limit
        r5 = ttk.Frame(c3, style="Card.TFrame")
        r5.pack(fill="x", pady=(0, 4))
        ttk.Checkbutton(r5, text="Split output by row limit:",
                        variable=self.use_limit, style="TCheckbutton",
                        command=self._toggle_limit).pack(side="left")
        self.limit_entry = ttk.Entry(r5, textvariable=self.row_limit, width=10,
                                     style="TEntry", state="disabled")
        self.limit_entry.pack(side="left", padx=(6, 0), ipady=3)
        ttk.Label(c3,
                  text="When enabled, data is split into multiple files once the row count per file is reached. File names are auto-numbered.",
                  style="Muted.TLabel", wraplength=280).pack(anchor="w")

        # Run
        ttk.Button(parent, text="▶  Run Consolidation", style="Run.TButton",
                   command=self._run).pack(fill="x", pady=(4, 0))

    # ── Right panel ───────────────────────────────────────────────
    def _build_right(self, parent):
        top = ttk.Frame(parent, style="Card.TFrame", padding=12)
        top.pack(fill="both", expand=True)

        hdr = ttk.Frame(top, style="Card.TFrame")
        hdr.pack(fill="x", pady=(0, 8))
        ttk.Label(hdr, text="Consolidation Results", style="Head.TLabel").pack(side="left")

        # Stats bar
        sf = tk.Frame(top, bg=self.C["card"])
        sf.pack(fill="x", pady=(0, 8))
        self._s_input  = self._badge(sf, "Input Files",   "—", self.C["accent"])
        self._s_read   = self._badge(sf, "Files Read",    "—", self.C["pass"])
        self._s_rows   = self._badge(sf, "Total Rows",    "—", self.C["fg"])
        self._s_out    = self._badge(sf, "Output Files",  "—", self.C["warn"])
        self._s_errors = self._badge(sf, "Errors",        "—", self.C["fail"])

        # Output files table
        ttk.Label(top, text="Output Files", style="Head.TLabel").pack(anchor="w", pady=(0, 4))
        out_frame = tk.Frame(top, bg=self.C["card"])
        out_frame.pack(fill="x", pady=(0, 8))
        vsb1 = ttk.Scrollbar(out_frame)
        vsb1.pack(side="right", fill="y")
        self.out_tree = ttk.Treeview(out_frame, columns=("file", "rows", "path"),
                                     show="headings", height=4,
                                     yscrollcommand=vsb1.set)
        self.out_tree.pack(fill="x")
        vsb1.config(command=self.out_tree.yview)
        self.out_tree.heading("file", text="File Name")
        self.out_tree.heading("rows", text="Data Rows")
        self.out_tree.heading("path", text="Full Path")
        self.out_tree.column("file", width=200, anchor="w")
        self.out_tree.column("rows", width=90,  anchor="center")
        self.out_tree.column("path", width=350, anchor="w")

        # Errors table
        ttk.Label(top, text="Skipped / Errors", style="Head.TLabel").pack(anchor="w", pady=(0, 4))
        err_frame = tk.Frame(top, bg=self.C["card"])
        err_frame.pack(fill="x", pady=(0, 8))
        vsb2 = ttk.Scrollbar(err_frame)
        vsb2.pack(side="right", fill="y")
        self.err_tree = ttk.Treeview(err_frame, columns=("file", "reason"),
                                     show="headings", height=3,
                                     yscrollcommand=vsb2.set)
        self.err_tree.pack(fill="x")
        vsb2.config(command=self.err_tree.yview)
        self.err_tree.heading("file",   text="File Name")
        self.err_tree.heading("reason", text="Reason")
        self.err_tree.column("file",   width=200, anchor="w")
        self.err_tree.column("reason", width=450, anchor="w")
        self.err_tree.tag_configure("err", foreground=self.C["fail"])

        # Log
        ttk.Label(top, text="Processing Log", style="Head.TLabel").pack(anchor="w", pady=(0, 4))
        log_frame = tk.Frame(top, bg=self.C["card"])
        log_frame.pack(fill="both", expand=True)
        lsb = ttk.Scrollbar(log_frame)
        lsb.pack(side="right", fill="y")
        self.log_text = tk.Text(log_frame, bg="#252839", fg=self.C["fg"],
                                font=("Courier", 8), relief="flat", wrap="word",
                                yscrollcommand=lsb.set, state="disabled",
                                borderwidth=0, highlightthickness=0)
        self.log_text.pack(fill="both", expand=True)
        lsb.config(command=self.log_text.yview)

    def _badge(self, parent, label, value, color):
        f = tk.Frame(parent, bg=self.C["card"], padx=10, pady=4)
        f.pack(side="left", padx=(0, 10))
        tk.Label(f, text=label, bg=self.C["card"], fg=self.C["muted"],
                 font=("Arial", 8)).pack()
        lbl = tk.Label(f, text=value, bg=self.C["card"], fg=color,
                       font=("Arial", 15, "bold"))
        lbl.pack()
        return lbl

    # ── Input helpers ─────────────────────────────────────────────
    def _toggle_input(self):
        if self.input_mode.get() == "files":
            self._folder_frame.pack_forget()
            self._files_frame.pack(fill="x")
        else:
            self._files_frame.pack_forget()
            self._folder_frame.pack(fill="x")

    def _toggle_limit(self):
        self.limit_entry.config(
            state="normal" if self.use_limit.get() else "disabled"
        )

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Select Excel Files",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls"), ("All", "*.*")])
        for p in paths:
            if p not in self.input_paths:
                self.input_paths.append(p)
                self.files_lb.insert(tk.END, os.path.basename(p))

    def _remove_selected(self):
        sel = list(self.files_lb.curselection())
        for i in reversed(sel):
            self.files_lb.delete(i)
            self.input_paths.pop(i)

    def _clear_files(self):
        self.input_paths.clear()
        self.files_lb.delete(0, tk.END)

    def _browse_folder(self):
        p = filedialog.askdirectory(title="Select Folder Containing Excel Files")
        if p:
            self.input_folder.set(p)

    def _browse_output_dir(self):
        p = filedialog.askdirectory(title="Select Output Folder")
        if p:
            self.output_dir.set(p)

    # ── Log helpers ───────────────────────────────────────────────
    def _log(self, msg):
        self.log_text.config(state="normal")
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.config(state="disabled")
        self.update_idletasks()

    def _clear_log(self):
        self.log_text.config(state="normal")
        self.log_text.delete("1.0", tk.END)
        self.log_text.config(state="disabled")

    # ── Run ───────────────────────────────────────────────────────
    def _run(self):
        # Collect file paths
        if self.input_mode.get() == "files":
            file_paths = list(self.input_paths)
        else:
            folder = self.input_folder.get().strip()
            if not folder or not os.path.isdir(folder):
                messagebox.showerror("Missing Input", "Please select a valid input folder.")
                return
            file_paths = []
            for ext in ("*.xlsx", "*.xlsm", "*.xls"):
                file_paths.extend(glob.glob(os.path.join(folder, ext)))
            file_paths.sort()

        if not file_paths:
            messagebox.showerror("No Files", "No Excel files found to consolidate.")
            return

        # Validate sheet position and header rows
        try:
            sheet_pos = int(self.sheet_position.get())
            assert sheet_pos >= 1
        except Exception:
            messagebox.showerror("Configuration", "Sheet position must be a positive integer.")
            return
        try:
            hdr_rows = int(self.header_rows.get())
            assert hdr_rows >= 0
        except Exception:
            messagebox.showerror("Configuration", "Header row count must be a non-negative integer.")
            return

        # Output name
        out_name = self.output_name.get().strip()
        if not out_name:
            messagebox.showerror("Configuration", "Please enter an output file name.")
            return
        # Strip .xlsx if user typed it
        if out_name.lower().endswith(".xlsx"):
            out_name = out_name[:-5]

        # Output directory
        out_dir = self.output_dir.get().strip()
        if not out_dir:
            out_dir = filedialog.askdirectory(title="Select Output Folder")
            if not out_dir:
                return
            self.output_dir.set(out_dir)
        if not os.path.isdir(out_dir):
            messagebox.showerror("Output Folder", "The specified output folder does not exist.")
            return

        # Row limit
        row_limit = None
        if self.use_limit.get():
            try:
                row_limit = int(self.row_limit.get())
                assert row_limit >= 1
            except Exception:
                messagebox.showerror("Configuration",
                                     "Row limit must be a positive integer when splitting is enabled.")
                return

        # Clear previous results
        self._clear_log()
        for tree in (self.out_tree, self.err_tree):
            for item in tree.get_children():
                tree.delete(item)
        for badge in (self._s_input, self._s_read, self._s_rows,
                      self._s_out, self._s_errors):
            badge.config(text="…")

        self._log(f"Starting consolidation — {len(file_paths)} file(s) queued")
        self._log(f"Sheet position: {sheet_pos}   Header rows: {hdr_rows}")
        if row_limit:
            self._log(f"Row limit per output file: {row_limit:,}")
        self._log("─" * 55)

        try:
            output_files, summary = consolidate(
                file_paths=file_paths,
                sheet_index=sheet_pos - 1,
                header_row_count=hdr_rows,
                output_basename=out_name,
                output_dir=out_dir,
                row_limit=row_limit,
                status_cb=self._log,
            )
        except RuntimeError as e:
            messagebox.showerror("Consolidation Error", str(e))
            self._log(f"ERROR: {e}")
            return

        # Summary
        self._log("─" * 55)
        self._log(f"Done.  {summary['total_data_rows']:,} data row(s) consolidated "
                  f"into {len(output_files)} output file(s).")

        self._s_input.config(text=str(summary["total_input_files"]))
        self._s_read.config(text=str(summary["files_read"]))
        self._s_rows.config(text=f"{summary['total_data_rows']:,}")
        self._s_out.config(text=str(len(output_files)))
        self._s_errors.config(text=str(len(summary["errors"])))

        # Populate output files table
        for op in output_files:
            # Count data rows in the file quickly via openpyxl
            try:
                wb_t = openpyxl.load_workbook(op, read_only=True)
                ws_t = wb_t.active
                data_row_count = ws_t.max_row - hdr_rows
                wb_t.close()
            except Exception:
                data_row_count = "—"
            self.out_tree.insert("", tk.END,
                                 values=(os.path.basename(op),
                                         f"{data_row_count:,}" if isinstance(data_row_count, int) else data_row_count,
                                         op))

        # Populate errors table
        for err in summary["errors"]:
            self.err_tree.insert("", tk.END,
                                 values=(err["file"], err["error"]),
                                 tags=("err",))

        if not summary["errors"]:
            messagebox.showinfo(
                "Consolidation Complete",
                f"{summary['total_data_rows']:,} data row(s) consolidated into "
                f"{len(output_files)} file(s).\n\nSaved to: {out_dir}"
            )


# ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = App()
    app.mainloop()