"""
Excel Header Checker — GUI Tool
Validates headers, sheet names, and cell formatting of one or more Excel files against a reference file.
v1.4 — Adds Excel formatting checks (font, bold, italic, color, fill, alignment, merged cells).

Run: python header_checker_v1.4.py
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import os
import glob
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
import csv
from datetime import datetime
from collections import defaultdict


# ═══════════════════════════════════════════════════════════════════════
# Formatting extraction helpers
# ═══════════════════════════════════════════════════════════════════════

def _font_to_dict(font):
    """Convert an openpyxl Font to a hashable dict for comparison."""
    if font is None:
        return {"name": None, "size": None, "bold": None, "italic": None, "color": None}
    color = None
    if font.color and font.color.rgb:
        color = str(font.color.rgb).lstrip("0") or "00000000"
    return {
        "name": font.name,
        "size": font.size,
        "bold": bool(font.bold),
        "italic": bool(font.italic),
        "color": color,
    }


def _fill_to_dict(fill):
    """Convert an openpyxl PatternFill to a hashable dict."""
    if fill is None or not isinstance(fill, PatternFill):
        return {"type": None, "fgColor": None, "bgColor": None}
    fg = None
    bg = None
    if fill.fgColor and fill.fgColor.rgb:
        fg = str(fill.fgColor.rgb).lstrip("0") or "00000000"
    if fill.bgColor and fill.bgColor.rgb:
        bg = str(fill.bgColor.rgb).lstrip("0") or "00000000"
    return {"type": fill.fill_type, "fgColor": fg, "bgColor": bg}


def _align_to_dict(alignment):
    """Convert an openpyxl Alignment to a hashable dict."""
    if alignment is None:
        alignment = Alignment()
    return {
        "horizontal": alignment.horizontal,
        "vertical": alignment.vertical,
        "wrap_text": bool(alignment.wrap_text),
    }


def _get_merged_cell_map(ws, header_rows):
    """
    Return a dict mapping (row, col) -> (merge_start_row, merge_start_col, merge_end_row, merge_end_col)
    for cells that belong to merged ranges within the header rows.
    """
    merged = {}
    for merge_range in ws.merged_cells.ranges:
        # Only consider merges that intersect the header rows
        if merge_range.min_row > header_rows:
            continue
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            if row > header_rows:
                break
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merged[(row, col)] = (
                    merge_range.min_row,
                    merge_range.min_col,
                    merge_range.max_row,
                    merge_range.max_col,
                )
    return merged


def read_headers_with_formatting(filepath, sheet_configs):
    """
    Read headers, sheet names, AND cell formatting from an Excel file.
    sheet_configs: list of dicts [{sheet_index (0-based), header_row_count}]
    Returns:
      {sheet_index: {
          name: str,
          headers: list[str],
          fonts: list[dict],
          fills: list[dict],
          alignments: list[dict],
          merged: set of (row, col) tuples,
          merged_ranges: dict mapping (row,col) -> merge bounds
      }}
    """
    wb = openpyxl.load_workbook(filepath, read_only=False, data_only=False)
    sheet_names = wb.sheetnames
    result = {}
    for cfg in sheet_configs:
        si = cfg["sheet_index"]
        hr = cfg["header_row_count"]
        if si >= len(sheet_names):
            result[si] = None
            continue
        ws = wb.worksheets[si]

        # Read cell values and formatting row by row
        rows_vals = []
        rows_fonts = []
        rows_fills = []
        rows_aligns = []
        max_cols = 0
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=hr)):
            vals = []
            fonts = []
            fills = []
            aligns = []
            for cell in row:
                vals.append(cell.value)
                fonts.append(_font_to_dict(cell.font))
                fills.append(_fill_to_dict(cell.fill))
                aligns.append(_align_to_dict(cell.alignment))
            rows_vals.append(vals)
            rows_fonts.append(fonts)
            rows_fills.append(fills)
            rows_aligns.append(aligns)
            max_cols = max(max_cols, len(vals))

        # Normalise all rows to max_cols width
        for lst in (rows_vals, rows_fonts, rows_fills, rows_aligns):
            for row in lst:
                while len(row) < max_cols:
                    if isinstance(row[-1], dict) if lst is rows_vals else False:
                        row.append(None)
                    else:
                        row.append(None if lst is rows_vals else
                                   _font_to_dict(Font()) if lst is rows_fonts else
                                   _fill_to_dict(PatternFill()) if lst is rows_fills else
                                   _align_to_dict(Alignment()))

        # Flatten multi-row headers (same logic as v1.3 for text)
        if len(rows_vals) == 0:
            headers = []
        elif len(rows_vals) == 1:
            headers = [str(c) if c is not None else "" for c in rows_vals[0]]
        else:
            headers = []
            for ci in range(max_cols):
                parts = []
                for r in rows_vals:
                    val = None
                    if ci < len(r):
                        val = r[ci]
                    if val is not None and str(val).strip() != "":
                        parts.append(str(val).strip())
                headers.append(" | ".join(parts) if parts else "")

        # For formatting, use the last row's formatting as the representative (closest to data)
        # OR create flattened formatting dicts — for simplicity we take formatting from
        # the first non-empty cell in each column across header rows
        flat_fonts = []
        flat_fills = []
        flat_aligns = []
        for ci in range(max_cols):
            # Use the LAST header row's formatting as the canonical formatting
            last_row_font = rows_fonts[-1][ci] if rows_fonts else _font_to_dict(Font())
            last_row_fill = rows_fills[-1][ci] if rows_fills else _fill_to_dict(PatternFill())
            last_row_align = rows_aligns[-1][ci] if rows_aligns else _align_to_dict(Alignment())
            flat_fonts.append(last_row_font)
            flat_fills.append(last_row_fill)
            flat_aligns.append(last_row_align)

        merged_cells = _get_merged_cell_map(ws, hr)

        result[si] = {
            "name": sheet_names[si],
            "headers": headers,
            "fonts": flat_fonts,
            "fills": flat_fills,
            "alignments": flat_aligns,
            "merged": merged_cells,
        }
    wb.close()
    return result


# ═══════════════════════════════════════════════════════════════════════
# Formatting comparison helpers
# ═══════════════════════════════════════════════════════════════════════

def _fmt_diff_label(prop, ref_val, chk_val):
    """Generate a human-readable label for a formatting mismatch."""
    return f"{prop}: expected '{ref_val}', found '{chk_val}'"


def _compare_font(ref_font, chk_font, col_pos):
    """Return list of mismatch strings for a single font."""
    mismatches = []
    for key, label in [("name", "Font"), ("size", "Size"), ("bold", "Bold"),
                        ("italic", "Italic"), ("color", "Font Color")]:
        rv = ref_font.get(key)
        cv = chk_font.get(key)
        if rv != cv:
            mismatches.append(_fmt_diff_label(label, rv, cv))
    return mismatches


def _compare_fill(ref_fill, chk_fill, col_pos):
    """Return list of mismatch strings for a single fill."""
    mismatches = []
    for key, label in [("fgColor", "Fill Color"), ("bgColor", "Background Color")]:
        rv = ref_fill.get(key)
        cv = chk_fill.get(key)
        if rv != cv:
            mismatches.append(_fmt_diff_label(label, rv, cv))
    return mismatches


def _compare_alignment(ref_align, chk_align, col_pos):
    """Return list of mismatch strings for a single alignment."""
    mismatches = []
    for key, label in [("horizontal", "Horizontal Align"), ("vertical", "Vertical Align"),
                        ("wrap_text", "Wrap Text")]:
        rv = ref_align.get(key)
        cv = chk_align.get(key)
        if rv != cv:
            mismatches.append(_fmt_diff_label(label, rv, cv))
    return mismatches


def _compare_merged(ref_merged, chk_merged, col_count):
    """
    Compare merged cell regions. Returns list of mismatch strings.
    ref_merged, chk_merged: dicts from _get_merged_cell_map
    """
    mismatches = []
    # Check columns that should be merged but aren't (or vice versa)
    ref_merged_cols = set()
    chk_merged_cols = set()
    for (r, c), bounds in ref_merged.items():
        if r == 1:  # first header row
            ref_merged_cols.add(c)
    for (r, c), bounds in chk_merged.items():
        if r == 1:
            chk_merged_cols.add(c)

    extra_merged = chk_merged_cols - ref_merged_cols
    missing_merged = ref_merged_cols - chk_merged_cols

    if missing_merged:
        col_labels = ", ".join(str(c) for c in sorted(missing_merged))
        mismatches.append(f"Merged Cells: expected merged columns {col_labels}, but they are not merged")

    if extra_merged:
        col_labels = ", ".join(str(c) for c in sorted(extra_merged))
        mismatches.append(f"Merged Cells: unexpected merged columns {col_labels}")

    return mismatches


def generate_formatting_remarks(ref_info, chk_info, sheet_configs):
    """
    Compare formatting (font, fill, alignment, merged cells) between reference and checked file.
    Returns list of formatting remark strings.
    """
    remarks = []
    for cfg in sheet_configs:
        si = cfg["sheet_index"]
        ref = ref_info.get(si)
        chk = chk_info.get(si)
        sheet_label = f'Sheet "{ref["name"]}"'

        if chk is None:
            continue  # structural remark already handled in text comparison

        ref_hdrs = ref["headers"]
        chk_hdrs = chk["headers"]

        # Strip trailing blanks to get meaningful column count
        def strip_trailing(lst):
            lst = list(lst)
            while lst and lst[-1].strip() == "":
                lst.pop()
            return lst

        ref_stripped = strip_trailing(ref_hdrs)
        chk_stripped = strip_trailing(chk_hdrs)
        col_count = min(len(ref_stripped), len(chk_stripped))
        if col_count == 0:
            continue

        # Per-column formatting check
        for ci in range(col_count):
            col_label = ci + 1
            ref_header_name = ref_stripped[ci].strip() if ci < len(ref_stripped) else f"Col {col_label}"
            col_id = f"{sheet_label}: Col {col_label} ('{ref_header_name}')"

            font_mismatches = _compare_font(
                ref["fonts"][ci] if ci < len(ref["fonts"]) else _font_to_dict(Font()),
                chk["fonts"][ci] if ci < len(chk["fonts"]) else _font_to_dict(Font()),
                col_label,
            )
            for m in font_mismatches:
                remarks.append(f"{col_id} — {m}")

            fill_mismatches = _compare_fill(
                ref["fills"][ci] if ci < len(ref["fills"]) else _fill_to_dict(PatternFill()),
                chk["fills"][ci] if ci < len(chk["fills"]) else _fill_to_dict(PatternFill()),
                col_label,
            )
            for m in fill_mismatches:
                remarks.append(f"{col_id} — {m}")

            align_mismatches = _compare_alignment(
                ref["alignments"][ci] if ci < len(ref["alignments"]) else _align_to_dict(Alignment()),
                chk["alignments"][ci] if ci < len(chk["alignments"]) else _align_to_dict(Alignment()),
                col_label,
            )
            for m in align_mismatches:
                remarks.append(f"{col_id} — {m}")

        # Merged cell comparison
        merge_mismatches = _compare_merged(ref["merged"], chk["merged"], col_count)
        for m in merge_mismatches:
            remarks.append(f"{sheet_label} — {m}")

    return remarks


# ═══════════════════════════════════════════════════════════════════════
# Core validation logic (text + optional formatting)
# ═══════════════════════════════════════════════════════════════════════

def read_headers_text(filepath, sheet_configs):
    """Read headers and sheet names (text only) — same as v1.3."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    result = {}
    for cfg in sheet_configs:
        si = cfg["sheet_index"]
        hr = cfg["header_row_count"]
        if si >= len(sheet_names):
            result[si] = None
            continue
        ws = wb.worksheets[si]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= hr:
                break
            rows.append(list(row))
        if len(rows) == 0:
            headers = []
        elif len(rows) == 1:
            headers = [str(c) if c is not None else "" for c in rows[0]]
        else:
            col_count = max(len(r) for r in rows)
            headers = []
            for ci in range(col_count):
                parts = []
                for r in rows:
                    val = r[ci] if ci < len(r) else None
                    if val is not None and str(val).strip() != "":
                        parts.append(str(val).strip())
                headers.append(" | ".join(parts) if parts else "")
        result[si] = {"name": sheet_names[si], "headers": headers}
    wb.close()
    return result


def generate_remarks(ref_info, chk_info, sheet_configs):
    """
    Compare checked file header text against reference.
    Same core logic as v1.3.
    """
    def strip_trailing(lst):
        lst = list(lst)
        while lst and lst[-1].strip() == "":
            lst.pop()
        return lst

    def find_extra_columns(ref_hdrs, chk_hdrs):
        ref_seq = [h.strip() for h in ref_hdrs]
        chk_seq = [h.strip() for h in chk_hdrs]
        ri = 0
        extras = []
        for ci, ch in enumerate(chk_seq):
            if ri < len(ref_seq) and ch == ref_seq[ri]:
                ri += 1
            else:
                extras.append((ci, ch))
        if ri == len(ref_seq):
            return extras
        return None

    remarks = []
    all_pass = True

    for cfg in sheet_configs:
        si = cfg["sheet_index"]
        ref = ref_info.get(si)
        chk = chk_info.get(si)
        sheet_label = f'Sheet "{ref["name"]}"'

        if chk is None:
            remarks.append(f"{sheet_label}: Sheet is missing from this file.")
            all_pass = False
            continue

        # Sheet name
        if ref["name"] != chk["name"]:
            remarks.append(
                f'{sheet_label}: Sheet name is incorrect — found "{chk["name"]}".')
            all_pass = False

        ref_hdrs = strip_trailing(ref["headers"])
        chk_hdrs = strip_trailing(chk["headers"])
        ref_count = len(ref_hdrs)
        chk_count = len(chk_hdrs)

        # Extra-column / shift detection
        if chk_count > ref_count:
            extras = find_extra_columns(ref_hdrs, chk_hdrs)
            if extras is not None:
                all_pass = False
                blank_extras = [(i, v) for i, v in extras if v == ""]
                named_extras = [(i, v) for i, v in extras if v != ""]
                if blank_extras:
                    positions = ", ".join(str(i + 1) for i, _ in blank_extras)
                    count = len(blank_extras)
                    remarks.append(
                        f"{sheet_label}: {count} blank extra column(s) inserted "
                        f"at position(s) {positions}. "
                        f"All required headers are present but shifted accordingly.")
                for i, v in named_extras:
                    remarks.append(
                        f'{sheet_label}: An extra column "{v}" is inserted at position {i + 1}. '
                        f"All required headers are present but shifted accordingly.")
                continue

        # Column count deficit
        if chk_count < ref_count:
            all_pass = False
            missing_count = ref_count - chk_count
            remarks.append(
                f"{sheet_label}: {missing_count} column(s) are missing "
                f"(expected {ref_count}, found {chk_count}).")

        # Per-column check
        chk_value_positions = defaultdict(list)
        for ci, ch in enumerate(chk_hdrs):
            chk_value_positions[ch.strip()].append(ci)

        check_up_to = min(ref_count, chk_count)
        for idx in range(check_up_to):
            ref_h = ref_hdrs[idx].strip()
            chk_h = chk_hdrs[idx].strip()
            if chk_h == ref_h:
                continue
            all_pass = False
            pos = idx + 1
            if chk_h == "":
                remarks.append(
                    f'{sheet_label}: Column "{ref_h}" (position {pos}) is blank.')
            else:
                remarks.append(
                    f'{sheet_label}: Column "{ref_h}" (position {pos}) is incorrect — '
                    f'found "{chk_h}".')

        # Tail columns fully absent
        for idx in range(check_up_to, ref_count):
            all_pass = False
            ref_h = ref_hdrs[idx].strip()
            remarks.append(
                f'{sheet_label}: Column "{ref_h}" (position {idx + 1}) is missing.')

    if all_pass:
        return ["All headers and sheet name(s) pass."]
    return remarks


def validate_files(ref_path, file_paths, sheet_configs, check_formatting=False):
    """
    Main validation entry point.
    Returns list of dicts: {file, status, remarks}
    """
    try:
        ref_info_text = read_headers_text(ref_path, sheet_configs)
    except Exception as e:
        raise RuntimeError(f"Failed to read reference file: {e}")

    ref_info_fmt = None
    if check_formatting:
        try:
            ref_info_fmt = read_headers_with_formatting(ref_path, sheet_configs)
        except Exception as e:
            raise RuntimeError(f"Failed to read reference file formatting: {e}")

    results = []
    for fp in file_paths:
        fname = os.path.basename(fp)
        try:
            chk_info_text = read_headers_text(fp, sheet_configs)
            text_remarks = generate_remarks(ref_info_text, chk_info_text, sheet_configs)

            if check_formatting and ref_info_fmt is not None:
                chk_info_fmt = read_headers_with_formatting(fp, sheet_configs)
                fmt_remarks = generate_formatting_remarks(ref_info_fmt, chk_info_fmt, sheet_configs)
            else:
                fmt_remarks = []

            all_remarks = text_remarks + fmt_remarks
            # Determine pass/fail: text must be all-pass AND no formatting mismatches
            text_pass = (len(text_remarks) == 1 and text_remarks[0].startswith("All headers"))
            fmt_pass = (len(fmt_remarks) == 0)
            passed = text_pass and fmt_pass

            results.append({
                "file": fname,
                "status": "Pass" if passed else "Fail",
                "remarks": "; ".join(all_remarks),
            })
        except Exception as e:
            results.append({
                "file": fname,
                "status": "Error",
                "remarks": f"Unable to process file: {e}",
            })
    return results


# ═══════════════════════════════════════════════════════════════════════
# GUI
# ═══════════════════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Excel Header Checker v1.4")
        self.geometry("1200x820")
        self.minsize(950, 680)
        self.configure(bg="#0f1117")

        self.ref_path = tk.StringVar()
        self.check_mode = tk.StringVar(value="files")
        self.check_paths = []
        self.check_folder = tk.StringVar()
        self.sheet_configs = []
        self.results = []
        self.check_formatting = tk.BooleanVar(value=True)

        self._build_styles()
        self._build_ui()

    def _build_styles(self):
        style = ttk.Style(self)
        style.theme_use("clam")

        BG = "#0f1117"
        CARD = "#1a1d27"
        ACCENT = "#4f8ef7"
        FG = "#e8eaf0"
        MUTED = "#6b7280"
        PASS_C = "#22c55e"
        FAIL_C = "#ef4444"
        ERR_C = "#f59e0b"

        style.configure("TFrame", background=BG)
        style.configure("Card.TFrame", background=CARD)
        style.configure("TLabel", background=BG, foreground=FG,
                        font=("Helvetica", 10))
        style.configure("Card.TLabel", background=CARD, foreground=FG,
                        font=("Helvetica", 10))
        style.configure("Title.TLabel", background=BG, foreground=FG,
                        font=("Helvetica", 18, "bold"))
        style.configure("Sub.TLabel", background=BG, foreground=MUTED,
                        font=("Helvetica", 10))
        style.configure("Head.TLabel", background=CARD, foreground=ACCENT,
                        font=("Helvetica", 11, "bold"))
        style.configure("TButton", background=ACCENT, foreground="#ffffff",
                        font=("Helvetica", 10, "bold"), borderwidth=0, padding=6)
        style.map("TButton",
                  background=[("active", "#6ba3ff"), ("disabled", "#2d3148")],
                  foreground=[("disabled", MUTED)])
        style.configure("Ghost.TButton", background=CARD, foreground=FG,
                        font=("Helvetica", 10), borderwidth=1, padding=5)
        style.map("Ghost.TButton", background=[("active", "#252839")])
        style.configure("Run.TButton", background="#22c55e", foreground="#000000",
                        font=("Helvetica", 11, "bold"), padding=8)
        style.map("Run.TButton", background=[("active", "#16a34a")])
        style.configure("TEntry", fieldbackground="#252839", foreground=FG,
                        insertcolor=FG, borderwidth=0)
        style.configure("TRadiobutton", background=BG, foreground=FG,
                        font=("Helvetica", 10))
        style.configure("TCheckbutton", background=BG, foreground=FG,
                        font=("Helvetica", 10))
        style.configure("Treeview", background=CARD, fieldbackground=CARD,
                        foreground=FG, font=("Helvetica", 9), rowheight=28,
                        borderwidth=0)
        style.configure("Treeview.Heading", background="#252839", foreground=ACCENT,
                        font=("Helvetica", 10, "bold"), borderwidth=0)
        style.map("Treeview", background=[("selected", "#2d3f6b")])
        style.configure("TSeparator", background="#252839")

        self.colors = {
            "bg": BG, "card": CARD, "accent": ACCENT, "fg": FG,
            "muted": MUTED, "pass": PASS_C, "fail": FAIL_C, "err": ERR_C,
        }

    def _build_ui(self):
        # Header
        header = tk.Frame(self, bg="#0d0f18", height=56)
        header.pack(fill="x")
        header.pack_propagate(False)
        tk.Label(header, text="⬡  Excel Header Checker v1.4",
                 bg="#0d0f18", fg=self.colors["accent"],
                 font=("Helvetica", 15, "bold")).pack(side="left", padx=20, pady=14)
        tk.Label(header, text="Header text + Excel formatting validation",
                 bg="#0d0f18", fg=self.colors["muted"],
                 font=("Helvetica", 9)).pack(side="left", pady=14)

        # Body
        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=16, pady=12)

        # Left panel
        left = ttk.Frame(body, width=360)
        left.pack(side="left", fill="y", padx=(0, 10))
        left.pack_propagate(False)
        self._build_left(left)

        # Right panel
        right = ttk.Frame(body)
        right.pack(side="left", fill="both", expand=True)
        self._build_right(right)

    def _build_left(self, parent):
        def card(p, title):
            f = ttk.Frame(p, style="Card.TFrame", padding=12)
            f.pack(fill="x", pady=(0, 10))
            ttk.Label(f, text=title, style="Head.TLabel").pack(anchor="w", pady=(0, 8))
            return f

        # 1. Reference file
        c1 = card(parent, "\u2460 Reference File")
        row = ttk.Frame(c1, style="Card.TFrame")
        row.pack(fill="x")
        ttk.Entry(row, textvariable=self.ref_path,
                  style="TEntry").pack(side="left", fill="x", expand=True, ipady=4)
        ttk.Button(row, text="Browse", style="Ghost.TButton",
                   command=self._browse_ref).pack(side="left", padx=(6, 0))

        # 2. Files to check
        c2 = card(parent, "\u2461 Files to Check")
        rb_row = ttk.Frame(c2, style="Card.TFrame")
        rb_row.pack(fill="x", pady=(0, 6))
        ttk.Radiobutton(rb_row, text="Select Files", variable=self.check_mode,
                        value="files", command=self._toggle_mode).pack(side="left", padx=(0, 12))
        ttk.Radiobutton(rb_row, text="Select Folder", variable=self.check_mode,
                        value="folder", command=self._toggle_mode).pack(side="left")

        # Files list
        self.files_frame = ttk.Frame(c2, style="Card.TFrame")
        self.files_frame.pack(fill="x")
        list_frame = tk.Frame(self.files_frame, bg=self.colors["card"])
        list_frame.pack(fill="x")
        sb = ttk.Scrollbar(list_frame)
        sb.pack(side="right", fill="y")
        self.files_lb = tk.Listbox(list_frame, height=5, bg="#252839",
                                   fg=self.colors["fg"], selectbackground=self.colors["accent"],
                                   font=("Helvetica", 9), relief="flat",
                                   yscrollcommand=sb.set, borderwidth=0,
                                   highlightthickness=0)
        self.files_lb.pack(fill="x")
        sb.config(command=self.files_lb.yview)
        btn_row = ttk.Frame(self.files_frame, style="Card.TFrame")
        btn_row.pack(fill="x", pady=(5, 0))
        ttk.Button(btn_row, text="+ Add Files", style="Ghost.TButton",
                   command=self._add_files).pack(side="left")
        ttk.Button(btn_row, text="Clear", style="Ghost.TButton",
                   command=self._clear_files).pack(side="left", padx=6)

        # Folder picker
        self.folder_frame = ttk.Frame(c2, style="Card.TFrame")
        folder_row = ttk.Frame(self.folder_frame, style="Card.TFrame")
        folder_row.pack(fill="x")
        ttk.Entry(folder_row, textvariable=self.check_folder,
                  style="TEntry").pack(side="left", fill="x", expand=True, ipady=4)
        ttk.Button(folder_row, text="Browse", style="Ghost.TButton",
                   command=self._browse_folder).pack(side="left", padx=(6, 0))
        self._toggle_mode()

        # 3. Formatting checkbox
        c3 = card(parent, "\u2462 Formatting Options")
        fmt_row = ttk.Frame(c3, style="Card.TFrame")
        fmt_row.pack(fill="x")
        ttk.Checkbutton(fmt_row, text="Check Excel formatting (font, color, fill, alignment, merged cells)",
                        variable=self.check_formatting).pack(anchor="w")
        ttk.Label(c3, text="When enabled, cell formatting in header rows is compared against the reference file.",
                  style="Card.TLabel", wraplength=310, font=("Helvetica", 8),
                  foreground=self.colors["muted"]).pack(anchor="w", pady=(4, 0))

        # 4. Sheet configuration
        c4 = card(parent, "\u2463 Sheet Configuration")
        ttk.Label(c4, text="Define which sheets to validate and how many header rows each has.",
                  style="Card.TLabel", wraplength=310, font=("Helvetica", 8),
                  foreground=self.colors["muted"]).pack(anchor="w", pady=(0, 8))

        self.sheet_rows_frame = ttk.Frame(c4, style="Card.TFrame")
        self.sheet_rows_frame.pack(fill="x", pady=(4, 0))
        self.sheet_row_widgets = []
        ttk.Button(c4, text="+ Add Sheet", style="Ghost.TButton",
                   command=self._add_sheet_row).pack(anchor="w", pady=(6, 0))
        self._add_sheet_row()

        # Run button
        ttk.Button(parent, text="\u25b6  Run Validation", style="Run.TButton",
                   command=self._run).pack(fill="x", pady=(4, 0))

    def _toggle_mode(self):
        if self.check_mode.get() == "files":
            self.folder_frame.pack_forget()
            self.files_frame.pack(fill="x")
        else:
            self.files_frame.pack_forget()
            self.folder_frame.pack(fill="x")

    def _add_sheet_row(self):
        idx = len(self.sheet_row_widgets)
        row = ttk.Frame(self.sheet_rows_frame, style="Card.TFrame")
        row.pack(fill="x", pady=2)
        ttk.Label(row, text=f"Sheet {idx+1}  Position:",
                  style="Card.TLabel", font=("Helvetica", 9)).pack(side="left")
        pos_var = tk.StringVar(value=str(idx + 1))
        ttk.Entry(row, textvariable=pos_var, width=4,
                  style="TEntry").pack(side="left", padx=(3, 8), ipady=3)
        ttk.Label(row, text="Header rows:", style="Card.TLabel",
                  font=("Helvetica", 9)).pack(side="left")
        hr_var = tk.StringVar(value="1")
        ttk.Entry(row, textvariable=hr_var, width=4,
                  style="TEntry").pack(side="left", padx=(3, 8), ipady=3)
        del_btn = ttk.Button(row, text="\u2715", style="Ghost.TButton",
                             command=lambda r=row, i=idx: self._remove_sheet_row(r, i))
        del_btn.pack(side="left")
        self.sheet_row_widgets.append({"frame": row, "pos": pos_var, "hr": hr_var})

    def _remove_sheet_row(self, frame, idx):
        frame.destroy()
        self.sheet_row_widgets = [w for w in self.sheet_row_widgets if w["frame"] != frame]

    def _browse_ref(self):
        path = filedialog.askopenfilename(
            title="Select Reference File",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls"), ("All", "*.*")])
        if path:
            self.ref_path.set(path)

    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="Select Files to Check",
            filetypes=[("Excel Files", "*.xlsx *.xlsm *.xls"), ("All", "*.*")])
        for p in paths:
            if p not in self.check_paths:
                self.check_paths.append(p)
                self.files_lb.insert(tk.END, os.path.basename(p))

    def _clear_files(self):
        self.check_paths = []
        self.files_lb.delete(0, tk.END)

    def _browse_folder(self):
        path = filedialog.askdirectory(title="Select Folder Containing Files")
        if path:
            self.check_folder.set(path)

    def _build_right(self, parent):
        top = ttk.Frame(parent, style="Card.TFrame", padding=12)
        top.pack(fill="both", expand=True)

        hdr = ttk.Frame(top, style="Card.TFrame")
        hdr.pack(fill="x", pady=(0, 10))
        ttk.Label(hdr, text="Validation Results", style="Head.TLabel").pack(side="left")

        # Export buttons
        exp = ttk.Frame(hdr, style="Card.TFrame")
        exp.pack(side="right")
        ttk.Button(exp, text="Export CSV", style="Ghost.TButton",
                   command=self._export_csv).pack(side="left", padx=4)
        ttk.Button(exp, text="Export Excel", style="Ghost.TButton",
                   command=self._export_excel).pack(side="left")

        # Stats bar
        self.stats_frame = tk.Frame(top, bg=self.colors["card"])
        self.stats_frame.pack(fill="x", pady=(0, 8))
        self.stat_total = self._stat_badge(self.stats_frame, "Total", "0", self.colors["muted"])
        self.stat_pass = self._stat_badge(self.stats_frame, "Pass", "0", self.colors["pass"])
        self.stat_fail = self._stat_badge(self.stats_frame, "Fail", "0", self.colors["fail"])
        self.stat_err = self._stat_badge(self.stats_frame, "Error", "0", self.colors["err"])

        # Treeview
        cols = ("file", "status", "remarks")
        tree_frame = tk.Frame(top, bg=self.colors["card"])
        tree_frame.pack(fill="both", expand=True)
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        vsb.pack(side="right", fill="y")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        hsb.pack(side="bottom", fill="x")
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings",
                                 yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tree.pack(fill="both", expand=True)
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)

        self.tree.heading("file", text="File Name")
        self.tree.heading("status", text="Status")
        self.tree.heading("remarks", text="Remarks")
        self.tree.column("file", width=220, minwidth=140, anchor="w")
        self.tree.column("status", width=80, minwidth=70, anchor="center")
        self.tree.column("remarks", width=650, minwidth=300, anchor="w")

        self.tree.tag_configure("pass", foreground=self.colors["pass"])
        self.tree.tag_configure("fail", foreground=self.colors["fail"])
        self.tree.tag_configure("err", foreground=self.colors["err"])

        # Detail pane
        detail_lbl = ttk.Label(top, text="Selected Row \u2014 Remarks Detail",
                               style="Head.TLabel")
        detail_lbl.pack(anchor="w", pady=(10, 4))
        self.detail_text = tk.Text(top, height=6, bg="#252839", fg=self.colors["fg"],
                                   font=("Helvetica", 9), relief="flat", wrap="word",
                                   state="disabled", borderwidth=0,
                                   highlightthickness=0, insertbackground=self.colors["fg"])
        self.detail_text.pack(fill="x")
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

    def _stat_badge(self, parent, label, value, color):
        f = tk.Frame(parent, bg=self.colors["card"], padx=12, pady=4)
        f.pack(side="left", padx=(0, 10))
        tk.Label(f, text=label, bg=self.colors["card"], fg=self.colors["muted"],
                 font=("Helvetica", 8)).pack()
        lbl = tk.Label(f, text=value, bg=self.colors["card"], fg=color,
                       font=("Helvetica", 16, "bold"))
        lbl.pack()
        return lbl

    def _on_select(self, _event):
        sel = self.tree.selection()
        if not sel:
            return
        item = self.tree.item(sel[0])
        remark = item["values"][2] if item["values"] else ""
        self.detail_text.config(state="normal")
        self.detail_text.delete("1.0", tk.END)
        for part in str(remark).split(";"):
            part = part.strip()
            if part:
                self.detail_text.insert(tk.END, f"\u2022 {part}\n")
        self.detail_text.config(state="disabled")

    def _run(self):
        # Validate inputs
        if not self.ref_path.get():
            messagebox.showerror("Missing Input", "Please select a reference file.")
            return

        if self.check_mode.get() == "files":
            file_paths = self.check_paths
        else:
            folder = self.check_folder.get()
            if not folder:
                messagebox.showerror("Missing Input", "Please select a folder.")
                return
            file_paths = []
            for ext in ("*.xlsx", "*.xlsm", "*.xls"):
                file_paths.extend(glob.glob(os.path.join(folder, ext)))

        if not file_paths:
            messagebox.showerror("No Files", "No Excel files found to check.")
            return

        if not self.sheet_row_widgets:
            messagebox.showerror("Configuration", "Please add at least one sheet to validate.")
            return

        sheet_configs = []
        for w in self.sheet_row_widgets:
            try:
                pos = int(w["pos"].get())
                hr = int(w["hr"].get())
                if pos < 1 or hr < 1:
                    raise ValueError
                sheet_configs.append({"sheet_index": pos - 1, "header_row_count": hr})
            except ValueError:
                messagebox.showerror("Configuration Error",
                                     "Sheet position and header row count must be positive integers.")
                return

        # Run
        try:
            results = validate_files(
                self.ref_path.get(), file_paths, sheet_configs,
                check_formatting=self.check_formatting.get(),
            )
        except RuntimeError as e:
            messagebox.showerror("Reference File Error", str(e))
            return

        self.results = results
        self._populate_tree(results)

    def _populate_tree(self, results):
        for item in self.tree.get_children():
            self.tree.delete(item)

        total = len(results)
        passed = sum(1 for r in results if r["status"] == "Pass")
        failed = sum(1 for r in results if r["status"] == "Fail")
        errs = sum(1 for r in results if r["status"] == "Error")

        self.stat_total.config(text=str(total))
        self.stat_pass.config(text=str(passed))
        self.stat_fail.config(text=str(failed))
        self.stat_err.config(text=str(errs))

        for r in results:
            tag = "pass" if r["status"] == "Pass" else (
                "err" if r["status"] == "Error" else "fail")
            status_display = ("\u2714 Pass" if r["status"] == "Pass" else
                              ("\u26a0 Error" if r["status"] == "Error" else "\u2716 Fail"))
            self.tree.insert("", tk.END,
                             values=(r["file"], status_display, r["remarks"]),
                             tags=(tag,))

    def _export_csv(self):
        if not self.results:
            messagebox.showinfo("No Results", "Run validation first.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV Files", "*.csv")],
            initialfile=f"header_check_{datetime.now():%Y%m%d_%H%M%S}.csv")
        if not path:
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["file", "status", "remarks"])
            writer.writeheader()
            writer.writerows(self.results)
        messagebox.showinfo("Exported", f"CSV saved to:\n{path}")

    def _export_excel(self):
        if not self.results:
            messagebox.showinfo("No Results", "Run validation first.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")],
            initialfile=f"header_check_{datetime.now():%Y%m%d_%H%M%S}.xlsx")
        if not path:
            return
        df = pd.DataFrame(self.results, columns=["file", "status", "remarks"])
        df.columns = ["File Name", "Status", "Remarks"]
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Validation Results")
            ws = writer.sheets["Validation Results"]
            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 12
            ws.column_dimensions["C"].width = 100
            from openpyxl.styles import Font as OFont, PatternFill as OPatternFill
            header_fill = OPatternFill("solid", fgColor="1a1d27")
            for cell in ws[1]:
                cell.font = OFont(bold=True, color="4F8EF7")
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for row in ws.iter_rows(min_row=2):
                status = row[1].value or ""
                color = "22c55e" if status == "Pass" else (
                    "f59e0b" if status == "Error" else "ef4444")
                row[1].font = OFont(color=color, bold=True)
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
        messagebox.showinfo("Exported", f"Excel saved to:\n{path}")


# ═══════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = App()
    app.mainloop()
