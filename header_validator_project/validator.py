from openpyxl import load_workbook

def get_headers(ws, header_row):
    return [cell.value if cell.value is not None else "" for cell in ws[header_row]]

def validate_file(ref_file, target_file, sheet_config):
    results = []
    ref_wb = load_workbook(ref_file, read_only=True, data_only=True)
    tgt_wb = load_workbook(target_file, read_only=True, data_only=True)

    for sheet_name, header_row in sheet_config.items():
        if sheet_name not in tgt_wb.sheetnames:
            results.append(("FAIL", sheet_name, f"Required worksheet is missing: {sheet_name}"))
            continue

        ref_ws = ref_wb[sheet_name]
        tgt_ws = tgt_wb[sheet_name]

        ref_headers = get_headers(ref_ws, header_row)
        tgt_headers = get_headers(tgt_ws, header_row)

        blanks = [i+1 for i,v in enumerate(tgt_headers) if str(v).strip()==""]
        if blanks:
            results.append(("FAIL", sheet_name,
                            f"Blank header column(s) detected at column position(s): {', '.join(map(str, blanks))}"))
            continue

        if len(ref_headers) != len(tgt_headers):
            missing = [h for h in ref_headers if h not in tgt_headers]
            extra = [h for h in tgt_headers if h not in ref_headers]
            msg = []
            if missing:
                msg.append("Required column(s) are missing: " + ", ".join(map(str, missing)))
            if extra:
                msg.append("Unexpected additional column(s) detected: " + ", ".join(map(str, extra)))
            results.append(("FAIL", sheet_name, " | ".join(msg)))
            continue

        mismatch = None
        for idx, (r, t) in enumerate(zip(ref_headers, tgt_headers), start=1):
            if r != t:
                mismatch = f"Header name mismatch detected at column position {idx}. Expected '{r}' but found '{t}'."
                break

        if mismatch:
            results.append(("FAIL", sheet_name, mismatch))
        else:
            results.append(("PASS", sheet_name, "All sheet names and headers passed validation."))

    return results
