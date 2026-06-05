import openpyxl
from openpyxl.styles import Font, PatternFill
from pathlib import Path

# Define headers from the image
headers = ["BreachID", "DocumentID", "Custodian", "File Extension", "File Name", 
           "Category", "Full Name", "First", "Middle", "Last", "Suffix", "SSN", "DOB"]

# Define sheet names
sheet_names = ["Business", "Individual"]

# Create output folder if it doesn't exist
output_folder = Path("Output_Folder")
output_folder.mkdir(exist_ok=True)

# Generate 10 Excel files
for i in range(1, 11):
    # Create a new workbook
    wb = openpyxl.Workbook()
    
    # Remove the default sheet
    wb.remove(wb.active)
    
    # Create two sheets with headers
    for sheet_name in sheet_names:
        ws = wb.create_sheet(sheet_name)
        
        # Add headers in the first row
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            
            # Style the header row
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, start=1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = len(header) + 2
    
    # Save the workbook
    file_name = f"PSA_{i}.xlsx"
    file_path = output_folder / file_name
    wb.save(file_path)
    print(f"Created: {file_path}")

print(f"\nSuccessfully generated 10 Excel template files in Output_Folder/")
